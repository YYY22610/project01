"""Admin router: dashboard, participants, submissions, scores, configs, exports."""
import io
import csv
import json
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_, cast, String
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import get_current_admin
from app.models.user import User, UserStatus, ExperimentGroup
from app.models.behavior_log import BehaviorLog, ActionType
from app.models.task_submission import TaskSubmission
from app.models.reminder import Reminder
from app.models.questionnaire import QuestionnaireItem, QuestionnaireResponse
from app.models.admin_score import AdminScore, compute_total
from app.models.chat_message import ChatMessage
from app.models.system_config import SystemConfig
from app.models.admin_user import AdminUser
from app.schemas.auth import AdminLoginRequest, AdminTokenResponse
from app.schemas.common import DashboardStats, AdminScoreRequest, SystemConfigUpdate, OpenClawStatus
from app.schemas.questionnaire import (
    QuestionnaireItemAdminCreate,
    QuestionnaireItemAdminUpdate,
)
from app.services.auth_service import verify_password, create_admin_token

router = APIRouter()


@router.post("/login", response_model=AdminTokenResponse)
async def admin_login(req: AdminLoginRequest, db: AsyncSession = Depends(get_db)):
    """Admin login."""
    result = await db.execute(select(AdminUser).where(AdminUser.username == req.username))
    admin = result.scalar_one_or_none()

    if not admin or not verify_password(req.password, admin.password_hash):
        raise HTTPException(status_code=401, detail="管理员账号或密码错误")

    token = create_admin_token(admin.username)
    return AdminTokenResponse(access_token=token, username=admin.username)


@router.get("/dashboard", response_model=DashboardStats)
async def dashboard(admin: dict = Depends(get_current_admin), db: AsyncSession = Depends(get_db)):
    """Dashboard statistics — real numbers for the monitoring board."""
    total_row = (await db.execute(select(func.count(User.id)))).scalar()
    total = total_row or 0

    h_count = (await db.execute(
        select(func.count(User.id)).where(cast(User.group, String) == ExperimentGroup.H.value)
    )).scalar() or 0
    soa_count = (await db.execute(
        select(func.count(User.id)).where(cast(User.group, String) == ExperimentGroup.SOA.value)
    )).scalar() or 0
    moa_count = (await db.execute(
        select(func.count(User.id)).where(cast(User.group, String) == ExperimentGroup.MOA.value)
    )).scalar() or 0

    completed = (await db.execute(
        select(func.count(User.id)).where(cast(User.status, String) == UserStatus.QUESTIONNAIRE_COMPLETED.value)
    )).scalar() or 0
    completion_rate = round(completed / total * 100, 1) if total else 0.0

    # Average + stddev of task duration (used for abnormal detection, 3-sigma rule)
    avg_dur = (await db.execute(select(func.avg(TaskSubmission.duration_ms)))).scalar()
    avg_duration_ms = float(avg_dur) if avg_dur else None
    std_dur = (await db.execute(select(func.stddev(TaskSubmission.duration_ms)))).scalar()
    abnormal_count = 0
    if avg_dur is not None and std_dur is not None and std_dur > 0:
        threshold = float(avg_dur) + 3 * float(std_dur)
        abnormal_count = (await db.execute(
            select(func.count(TaskSubmission.id)).where(TaskSubmission.duration_ms > threshold)
        )).scalar() or 0

    # Recent registrations (last 5) with a coarse progress percentage
    progress_map = {
        UserStatus.REGISTERED: 10,
        UserStatus.CONSENTED: 25,
        UserStatus.DEMO_COMPLETED: 40,
        UserStatus.TASK_IN_PROGRESS: 65,
        UserStatus.TASK_COMPLETED: 85,
        UserStatus.QUESTIONNAIRE_COMPLETED: 100,
    }
    recent_res = await db.execute(select(User).order_by(User.created_at.desc()).limit(5))
    recent_users = recent_res.scalars().all()
    recent_registrations = [
        {
            "email": u.email,
            "created_at": u.created_at.isoformat() if u.created_at else None,
            "status": u.status.value,
            "progress": progress_map.get(u.status, 0),
        }
        for u in recent_users
    ]

    # Experiment status + API status (heuristic from config)
    exp_cfg = (await db.execute(
        select(SystemConfig).where(SystemConfig.key == "experiment_status")
    )).scalar_one_or_none()
    experiment_status = exp_cfg.value if exp_cfg else ("running" if total > 0 else "stopped")

    llm_cfg = (await db.execute(
        select(SystemConfig).where(SystemConfig.key == "llm_api_key")
    )).scalar_one_or_none()
    email_cfg = (await db.execute(
        select(SystemConfig).where(SystemConfig.key == "smtp_host")
    )).scalar_one_or_none()
    api_status = {
        "search": "ok",
        "llm": "ok" if (llm_cfg and llm_cfg.value) else "unknown",
        "email": "ok" if (email_cfg and email_cfg.value) else "unknown",
    }

    target_cfg = (await db.execute(
        select(SystemConfig).where(SystemConfig.key == "target_sample_size")
    )).scalar_one_or_none()
    target = int(target_cfg.value) if target_cfg else 100

    # OpenClaw (AI agent service) runtime monitoring — derived from agent logs
    paused_cfg = (await db.execute(
        select(SystemConfig).where(SystemConfig.key == "agent_service_paused")
    )).scalar_one_or_none()
    paused = bool(paused_cfg and paused_cfg.value == "true")

    agent_calls = (await db.execute(
        select(func.count(BehaviorLog.id)).where(
            cast(BehaviorLog.action_type, String).in_([ActionType.AGENT_MESSAGE.value, ActionType.AGENT_TOOL_CALL.value])
        )
    )).scalar() or 0
    agent_success = (await db.execute(
        select(func.count(BehaviorLog.id)).where(
            cast(BehaviorLog.action_type, String).in_([ActionType.AGENT_MESSAGE.value, ActionType.AGENT_TOOL_CALL.value]),
            BehaviorLog.is_success == True,
        )
    )).scalar() or 0
    agent_latency = (await db.execute(
        select(func.avg(BehaviorLog.request_latency_ms)).where(
            cast(BehaviorLog.action_type, String).in_([ActionType.AGENT_MESSAGE.value, ActionType.AGENT_TOOL_CALL.value]),
            BehaviorLog.request_latency_ms.isnot(None),
        )
    )).scalar()
    agent_fail_24h = (await db.execute(
        select(func.count(BehaviorLog.id)).where(
            cast(BehaviorLog.action_type, String).in_([ActionType.AGENT_MESSAGE.value, ActionType.AGENT_TOOL_CALL.value]),
            BehaviorLog.is_success == False,
            BehaviorLog.timestamp >= datetime.now(timezone.utc) - timedelta(hours=24),
        )
    )).scalar() or 0

    if agent_calls == 0:
        oc_status = "paused" if paused else "unknown"
    elif agent_fail_24h > 0 or (agent_success / agent_calls) < 0.8:
        oc_status = "degraded"
    else:
        oc_status = "ok"
    if paused:
        oc_status = "paused"

    openclaw_status = OpenClawStatus(
        status=oc_status,
        paused=paused,
        total_calls=agent_calls,
        success_rate=round(agent_success / agent_calls, 3) if agent_calls else None,
        avg_latency_ms=round(float(agent_latency), 1) if agent_latency else None,
        recent_failures=agent_fail_24h,
    )

    return DashboardStats(
        total_participants=total,
        group_distribution={"H": h_count, "SOA": soa_count, "MOA": moa_count},
        experiment_status=experiment_status,
        completion_rate=completion_rate,
        avg_duration_ms=avg_duration_ms,
        api_status=api_status,
        openclaw_status=openclaw_status,
        recent_registrations=recent_registrations,
        abnormal_count=abnormal_count,
        target_sample=target,
    )


async def _compute_openclaw_status(db: AsyncSession) -> OpenClawStatus:
    """Shared helper: derive OpenClaw runtime status from agent behavior logs."""
    paused_cfg = (await db.execute(
        select(SystemConfig).where(SystemConfig.key == "agent_service_paused")
    )).scalar_one_or_none()
    paused = bool(paused_cfg and paused_cfg.value == "true")

    agent_calls = (await db.execute(
        select(func.count(BehaviorLog.id)).where(
            cast(BehaviorLog.action_type, String).in_([ActionType.AGENT_MESSAGE.value, ActionType.AGENT_TOOL_CALL.value])
        )
    )).scalar() or 0
    agent_success = (await db.execute(
        select(func.count(BehaviorLog.id)).where(
            cast(BehaviorLog.action_type, String).in_([ActionType.AGENT_MESSAGE.value, ActionType.AGENT_TOOL_CALL.value]),
            BehaviorLog.is_success == True,
        )
    )).scalar() or 0
    agent_latency = (await db.execute(
        select(func.avg(BehaviorLog.request_latency_ms)).where(
            cast(BehaviorLog.action_type, String).in_([ActionType.AGENT_MESSAGE.value, ActionType.AGENT_TOOL_CALL.value]),
            BehaviorLog.request_latency_ms.isnot(None),
        )
    )).scalar()
    agent_fail_24h = (await db.execute(
        select(func.count(BehaviorLog.id)).where(
            cast(BehaviorLog.action_type, String).in_([ActionType.AGENT_MESSAGE.value, ActionType.AGENT_TOOL_CALL.value]),
            BehaviorLog.is_success == False,
            BehaviorLog.timestamp >= datetime.now(timezone.utc) - timedelta(hours=24),
        )
    )).scalar() or 0

    if agent_calls == 0:
        oc_status = "paused" if paused else "unknown"
    elif agent_fail_24h > 0 or (agent_success / agent_calls) < 0.8:
        oc_status = "degraded"
    else:
        oc_status = "ok"
    if paused:
        oc_status = "paused"

    return OpenClawStatus(
        status=oc_status,
        paused=paused,
        total_calls=agent_calls,
        success_rate=round(agent_success / agent_calls, 3) if agent_calls else None,
        avg_latency_ms=round(float(agent_latency), 1) if agent_latency else None,
        recent_failures=agent_fail_24h,
    )


@router.get("/openclaw/status")
async def openclaw_status(admin: dict = Depends(get_current_admin), db: AsyncSession = Depends(get_db)):
    """OpenClaw (AI agent service) runtime status."""
    return await _compute_openclaw_status(db)


@router.post("/openclaw/toggle")
async def openclaw_toggle(
    req: dict,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Pause / resume the AI agent service (experiment interruption switch)."""
    paused = bool(req.get("paused", False))
    cfg = (await db.execute(
        select(SystemConfig).where(SystemConfig.key == "agent_service_paused")
    )).scalar_one_or_none()
    if cfg:
        cfg.value = "true" if paused else "false"
    else:
        cfg = SystemConfig(key="agent_service_paused", value="true" if paused else "false",
                           description="AI助理服务是否暂停")
        db.add(cfg)
    await db.commit()
    return {"status": "ok", "paused": paused}


@router.get("/logs")
async def list_all_logs(
    page: int = 1,
    page_size: int = 50,
    action_type: str | None = None,
    search: str | None = None,
    user_id: str | None = None,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """List all behavior logs (admin view) with enriched fields and filters."""
    from sqlalchemy import or_

    conditions = []
    if action_type:
        conditions.append(cast(BehaviorLog.action_type, String) == action_type)
    if user_id:
        conditions.append(BehaviorLog.user_id == user_id)
    if search:
        like = f"%{search}%"
        conditions.append(
            or_(
                BehaviorLog.input_content.ilike(like),
                BehaviorLog.ai_response.ilike(like),
                User.email.ilike(like),
            )
        )

    count_result = await db.execute(
        select(func.count(BehaviorLog.id))
        .join(User, BehaviorLog.user_id == User.id)
        .where(*(conditions if conditions else [True]))
    )
    total = count_result.scalar()

    result = await db.execute(
        select(BehaviorLog, User)
        .join(User, BehaviorLog.user_id == User.id)
        .where(*(conditions if conditions else [True]))
        .order_by(BehaviorLog.timestamp.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    rows = result.all()

    def serialize(log: BehaviorLog, user: User) -> dict:
        return {
            "id": str(log.id),
            "user_id": str(log.user_id),
            "user_email": user.email,
            "group": log.group,
            "action_type": log.action_type.value,
            "action_target": log.action_target,
            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
            "page_path": log.page_path,
            "input_content": log.input_content,
            "ai_response": log.ai_response,
            "agent_id": log.agent_id,
            "session_id": log.session_id,
            "request_latency_ms": log.request_latency_ms,
            "is_success": log.is_success,
            "error_detail": log.error_detail,
            "phase": log.phase,
            "user_action_on_ai": log.user_action_on_ai,
            "manual_edit_count": log.manual_edit_count,
            "results_viewed": log.results_viewed,
            "clicked_item_id": log.clicked_item_id,
            "ai_suggestion_id": log.ai_suggestion_id,
            "ai_suggestion_type": log.ai_suggestion_type,
            "ai_interaction_rounds": log.ai_interaction_rounds,
            "final_plan_submit_time": log.final_plan_submit_time.isoformat() if log.final_plan_submit_time else None,
            "extra_data": log.extra_data,
        }

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "logs": [serialize(log, user) for log, user in rows],
    }


@router.get("/participants")
async def list_participants(
    page: int = 1,
    page_size: int = 20,
    group: str | None = None,
    status: str | None = None,
    search: str | None = None,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """List all participants with filters."""
    query = select(User)
    count_query = select(func.count(User.id))

    if group:
        query = query.where(cast(User.group, String) == group)
        count_query = count_query.where(cast(User.group, String) == group)
    if status:
        query = query.where(cast(User.status, String) == status)
        count_query = count_query.where(cast(User.status, String) == status)
    if search:
        query = query.where(User.email.ilike(f"%{search}%"))
        count_query = count_query.where(User.email.ilike(f"%{search}%"))

    total = (await db.execute(count_query)).scalar()
    result = await db.execute(
        query.order_by(User.created_at.desc())
        .offset((page - 1) * page_size).limit(page_size)
    )
    users = result.scalars().all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "participants": [
            {
                "id": str(u.id),
                "email": u.email,
                "group": u.group.value if u.group else None,
                "status": u.status.value,
                "age": u.age,
                "gender": u.gender,
                "created_at": u.created_at.isoformat() if u.created_at else None,
                "task_start_time": u.task_start_time.isoformat() if u.task_start_time else None,
                "task_end_time": u.task_end_time.isoformat() if u.task_end_time else None,
            }
            for u in users
        ],
    }


@router.get("/participants/{user_id}")
async def get_participant_detail(
    user_id: str,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Get participant detail with logs, submission, and scores."""
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="参与者不存在")

    sub_result = await db.execute(select(TaskSubmission).where(TaskSubmission.user_id == user_id))
    submission = sub_result.scalar_one_or_none()

    score_result = await db.execute(select(AdminScore).where(AdminScore.user_id == user_id))
    scores = score_result.scalars().all()

    reminder_result = await db.execute(
        select(Reminder)
        .where(Reminder.user_id == user_id)
        .order_by(Reminder.reminder_datetime)
    )
    reminders = reminder_result.scalars().all()

    log_result = await db.execute(
        select(BehaviorLog).where(BehaviorLog.user_id == user_id)
        .order_by(BehaviorLog.timestamp.desc()).limit(100)
    )
    logs = log_result.scalars().all()

    # 文档下载文件名（管理员可下载参与者提交的 Word）
    docx_file_name = None
    if submission and submission.docx_file_path:
        import os
        docx_file_name = os.path.basename(submission.docx_file_path)

    return {
        "user": {
            "id": str(user.id), "email": user.email,
            "group": user.group.value if user.group else None,
            "status": user.status.value,
            "age": user.age, "gender": user.gender,
            "education": user.education, "tech_frequency": user.tech_frequency,
            "ai_experience": user.ai_experience,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "task_start_time": user.task_start_time.isoformat() if user.task_start_time else None,
            "task_end_time": user.task_end_time.isoformat() if user.task_end_time else None,
        },
        "submission": {
            "task1_search": submission.task1_search if submission else False,
            "task2_document": submission.task2_document if submission else False,
            "task3_reminder": submission.task3_reminder if submission else False,
            "task4_email": submission.task4_email if submission else False,
            "docx_file_path": submission.docx_file_path if submission else None,
            "docx_file_name": docx_file_name,
            "email_status": submission.email_status if submission else None,
            "email_recipient": submission.email_recipient if submission else None,
            "duration_ms": submission.duration_ms if submission else None,
            "submitted_at": submission.submitted_at.isoformat() if submission and submission.submitted_at else None,
        } if submission else None,
        "reminders": [
            {
                "id": str(r.id),
                "reminder_datetime": r.reminder_datetime.isoformat() if r.reminder_datetime else None,
                "content": r.content or "",
                "is_set": r.is_set,
            }
            for r in reminders
        ],
        "scores": [
            {
                "scenic_score": s.scenic_score,
                "historic_score": s.historic_score,
                "rarity_score": s.rarity_score,
                "scale_score": s.scale_score,
                "integrity_score": s.integrity_score,
                "fame_score": s.fame_score,
                "season_score": s.season_score,
                "eco_score": s.eco_score,
                "total_score": s.total_score,
                "quality_score": s.quality_score,
                "reminder_correct": s.reminder_correct,
                "notes": s.notes,
            }
            for s in scores
        ],
        "recent_logs": [
            {
                "action_type": log.action_type.value,
                "action_target": log.action_target,
                "timestamp": log.timestamp.isoformat() if log.timestamp else None,
                "input_content": log.input_content,
                "agent_id": log.agent_id,
                "request_latency_ms": log.request_latency_ms,
                "is_success": log.is_success,
                "user_action_on_ai": log.user_action_on_ai,
                "phase": log.phase,
                "results_viewed": log.results_viewed,
            }
            for log in logs[:20]
        ],
    }


@router.delete("/participants/{user_id}")
async def delete_participant(
    user_id: str,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Delete a participant and all associated data (logs, submissions, reminders, scores, responses, chat messages)."""
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="参与者不存在")

    # 级联删除关联数据，避免外键约束报错
    for model in (BehaviorLog, TaskSubmission, Reminder, QuestionnaireResponse, AdminScore, ChatMessage):
        await db.execute(model.__table__.delete().where(model.user_id == user_id))

    await db.delete(user)
    await db.commit()
    return {"success": True, "deleted_id": user_id}


@router.get("/participants/{user_id}/docx")
async def download_participant_docx(
    user_id: str,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Admin download of a participant's generated Word document (需求5.1.1/5.2)."""
    import os
    from fastapi.responses import FileResponse

    result = await db.execute(select(TaskSubmission).where(TaskSubmission.user_id == user_id))
    submission = result.scalar_one_or_none()
    if not submission or not submission.docx_file_path or not os.path.exists(submission.docx_file_path):
        raise HTTPException(status_code=404, detail="文档不存在")

    file_path = submission.docx_file_path
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=os.path.basename(file_path),
    )


@router.get("/submissions")
async def list_submissions(
    page: int = 1,
    page_size: int = 20,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """List all task submissions for scoring."""
    scored_subq = (
        select(func.count(AdminScore.id))
        .where(AdminScore.user_id == TaskSubmission.user_id)
        .correlate(TaskSubmission)
        .scalar_subquery()
    )
    total_subq = (
        select(func.max(AdminScore.total_score))
        .where(AdminScore.user_id == TaskSubmission.user_id)
        .correlate(TaskSubmission)
        .scalar_subquery()
    )
    result = await db.execute(
        select(TaskSubmission, User, scored_subq, total_subq)
        .join(User, TaskSubmission.user_id == User.id)
        .order_by(User.created_at.desc())
        .offset((page - 1) * page_size).limit(page_size)
    )
    rows = result.all()

    return {
        "page": page,
        "page_size": page_size,
        "submissions": [
            {
                "id": str(sub.id),
                "user_id": str(user.id),
                "email": user.email,
                "group": user.group.value if user.group else None,
                "task1_search": sub.task1_search,
                "task2_document": sub.task2_document,
                "task3_reminder": sub.task3_reminder,
                "task4_email": sub.task4_email,
                "duration_ms": sub.duration_ms,
                "submitted_at": sub.submitted_at.isoformat() if sub.submitted_at else None,
                "docx_file_path": sub.docx_file_path,
                "scored": (scored or 0) > 0,
                "total_score": tot,
            }
            for sub, user, scored, tot in rows
        ],
    }


@router.post("/scores/{user_id}")
async def set_score(
    user_id: str,
    req: AdminScoreRequest,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Set (upsert) the 8-factor researcher score for a participant."""
    total = compute_total(
        req.scenic_score, req.historic_score, req.rarity_score, req.scale_score,
        req.integrity_score, req.fame_score, req.season_score, req.eco_score,
    )
    existing = (await db.execute(
        select(AdminScore).where(AdminScore.user_id == user_id)
    )).scalar_one_or_none()

    if existing:
        existing.scenic_score = req.scenic_score
        existing.historic_score = req.historic_score
        existing.rarity_score = req.rarity_score
        existing.scale_score = req.scale_score
        existing.integrity_score = req.integrity_score
        existing.fame_score = req.fame_score
        existing.season_score = req.season_score
        existing.eco_score = req.eco_score
        existing.total_score = total
        existing.quality_score = req.quality_score
        existing.reminder_correct = bool(req.reminder_correct or False)
        existing.notes = req.notes
        existing.scored_by = admin.get("sub", "admin")
        existing.updated_at = datetime.now()
        score = existing
    else:
        score = AdminScore(
            user_id=user_id,
            scenic_score=req.scenic_score,
            historic_score=req.historic_score,
            rarity_score=req.rarity_score,
            scale_score=req.scale_score,
            integrity_score=req.integrity_score,
            fame_score=req.fame_score,
            season_score=req.season_score,
            eco_score=req.eco_score,
            total_score=total,
            quality_score=req.quality_score,
            reminder_correct=bool(req.reminder_correct or False),
            notes=req.notes,
            scored_by=admin.get("sub", "admin"),
        )
        db.add(score)

    await db.commit()
    await db.refresh(score)
    return {"status": "ok", "score_id": str(score.id), "total_score": total}


@router.get("/config")
async def get_configs(admin: dict = Depends(get_current_admin), db: AsyncSession = Depends(get_db)):
    """Get all system configs."""
    result = await db.execute(select(SystemConfig))
    configs = result.scalars().all()
    return {c.key: c.value for c in configs}


@router.put("/config")
async def update_config(
    req: SystemConfigUpdate,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Update a system config value."""
    result = await db.execute(select(SystemConfig).where(SystemConfig.key == req.key))
    config = result.scalar_one_or_none()
    if config:
        config.value = req.value
    else:
        config = SystemConfig(key=req.key, value=req.value)
        db.add(config)
    await db.commit()
    return {"status": "ok"}


def _question_type_to_frontend(item: QuestionnaireItem) -> str:
    """Map DB question_type + scale_level to frontend type tokens."""
    if item.question_type.value == "likert":
        return f"likert{item.scale_level or 5}"
    if item.question_type.value == "choice":
        # DB does not distinguish single/multiple; prefer single_choice when options exist.
        return "single_choice" if item.options else "choice"
    return item.question_type.value


def _frontend_type_to_db(payload_type: str, scale_level: int | None = None) -> tuple:
    """Return (QuestionType, scale_level) from frontend type token."""
    from app.models.questionnaire import QuestionType

    if payload_type == "likert5":
        return QuestionType.likert, 5
    if payload_type == "likert7":
        return QuestionType.likert, 7
    if payload_type in ("single_choice", "multiple_choice"):
        return QuestionType.choice, scale_level or 5
    if payload_type == "text":
        return QuestionType.text, scale_level or 5
    # Fallback
    return QuestionType.likert, scale_level or 5


def _item_to_admin_response(item: QuestionnaireItem) -> dict:
    return {
        "id": str(item.id),
        "construct": item.construct,
        "text": item.question_text,
        "type": _question_type_to_frontend(item),
        "options": item.options,
        "scale_level": item.scale_level,
        "sort_order": item.sort_order,
        "is_active": item.is_active,
        "applicable_groups": item.applicable_groups,
    }


@router.get("/questionnaire-config")
async def get_questionnaire_config(admin: dict = Depends(get_current_admin), db: AsyncSession = Depends(get_db)):
    """Get all questionnaire items for admin management (frontend-friendly format)."""
    result = await db.execute(
        select(QuestionnaireItem).order_by(QuestionnaireItem.sort_order)
    )
    items = result.scalars().all()
    return [_item_to_admin_response(item) for item in items]


@router.post("/questionnaire-config")
async def create_questionnaire_item(
    req: QuestionnaireItemAdminCreate,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Create a new questionnaire item."""
    question_type, scale_level = _frontend_type_to_db(req.type, req.scale_level)

    # Auto-assign sort_order to end if not provided
    sort_order = req.sort_order
    if sort_order == 0:
        max_order = (await db.execute(
            select(func.coalesce(func.max(QuestionnaireItem.sort_order), 0))
        )).scalar()
        sort_order = int(max_order) + 1

    item = QuestionnaireItem(
        construct=req.construct,
        question_text=req.text,
        question_type=question_type,
        options=req.options,
        scale_level=scale_level,
        sort_order=sort_order,
        is_active=req.is_active,
        applicable_groups=req.applicable_groups,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return _item_to_admin_response(item)


@router.put("/questionnaire-config/{item_id}")
async def update_questionnaire_item(
    item_id: str,
    req: QuestionnaireItemAdminUpdate,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Update an existing questionnaire item."""
    result = await db.execute(select(QuestionnaireItem).where(QuestionnaireItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="题项不存在")

    if req.construct is not None:
        item.construct = req.construct
    if req.text is not None:
        item.question_text = req.text
    if req.type is not None or req.scale_level is not None:
        question_type, scale_level = _frontend_type_to_db(
            req.type or _question_type_to_frontend(item),
            req.scale_level if req.scale_level is not None else item.scale_level,
        )
        item.question_type = question_type
        item.scale_level = scale_level
    if req.options is not None:
        item.options = req.options
    if req.sort_order is not None:
        item.sort_order = req.sort_order
    if req.is_active is not None:
        item.is_active = req.is_active
    if req.applicable_groups is not None:
        item.applicable_groups = req.applicable_groups

    await db.commit()
    await db.refresh(item)
    return _item_to_admin_response(item)


@router.delete("/questionnaire-config/{item_id}")
async def delete_questionnaire_item(
    item_id: str,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Delete a questionnaire item."""
    result = await db.execute(select(QuestionnaireItem).where(QuestionnaireItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="题项不存在")

    await db.delete(item)
    await db.commit()
    return {"status": "ok"}


@router.get("/export/participants")
async def export_participants(admin: dict = Depends(get_current_admin), db: AsyncSession = Depends(get_db)):
    """Export participant data as CSV."""
    result = await db.execute(select(User).order_by(User.created_at))
    users = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Email", "Group", "Status", "Age", "Gender", "Education",
                     "Tech_Frequency", "AI_Experience", "Created_At", "Task_Start", "Task_End"])
    for u in users:
        writer.writerow([
            str(u.id), u.email, u.group.value if u.group else "", u.status.value,
            u.age or "", u.gender or "", u.education or "",
            u.tech_frequency or "", u.ai_experience or "",
            u.created_at.isoformat() if u.created_at else "",
            u.task_start_time.isoformat() if u.task_start_time else "",
            u.task_end_time.isoformat() if u.task_end_time else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=participants_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


@router.get("/export/logs")
async def export_logs(admin: dict = Depends(get_current_admin), db: AsyncSession = Depends(get_db)):
    """Export all behavior logs as CSV."""
    result = await db.execute(
        select(BehaviorLog, User)
        .join(User, BehaviorLog.user_id == User.id)
        .order_by(BehaviorLog.timestamp)
    )
    rows = result.all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Log_ID", "User_ID", "Email", "Group", "Timestamp", "Action_Type",
                     "Action_Target", "Input_Content", "Page_Path", "Agent_ID",
                     "Latency_Ms", "Is_Success", "Error_Detail", "Phase",
                     "User_Action_On_AI", "Manual_Edit_Count", "Results_Viewed",
                     "Clicked_Item_ID", "AI_Suggestion_ID", "AI_Suggestion_Type",
                     "AI_Interaction_Rounds", "Final_Plan_Submit_Time"])
    for log, user in rows:
        writer.writerow([
            str(log.id), str(log.user_id), user.email,
            log.group or "", log.timestamp.isoformat() if log.timestamp else "",
            log.action_type.value, log.action_target or "",
            (log.input_content or "")[:200], log.page_path or "",
            log.agent_id or "",
            log.request_latency_ms if log.request_latency_ms is not None else "",
            log.is_success if log.is_success is not None else "",
            (log.error_detail or "")[:200], log.phase or "",
            log.user_action_on_ai or "", log.manual_edit_count if log.manual_edit_count is not None else "",
            log.results_viewed if log.results_viewed is not None else "",
            log.clicked_item_id or "", log.ai_suggestion_id or "",
            log.ai_suggestion_type or "", log.ai_interaction_rounds if log.ai_interaction_rounds is not None else "",
            log.final_plan_submit_time.isoformat() if log.final_plan_submit_time else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=logs_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


@router.get("/export/scores")
async def export_scores(admin: dict = Depends(get_current_admin), db: AsyncSession = Depends(get_db)):
    """Export all scores as CSV."""
    result = await db.execute(
        select(AdminScore, User)
        .join(User, AdminScore.user_id == User.id)
    )
    rows = result.all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Score_ID", "User_ID", "Email", "Group",
        "Scenic", "Historic", "Rarity", "Scale", "Integrity", "Fame", "Season",
        "Eco", "Total_Score", "Notes", "Scored_By", "Created_At",
    ])
    for score, user in rows:
        writer.writerow([
            str(score.id), str(score.user_id), user.email,
            user.group.value if user.group else "",
            score.scenic_score if score.scenic_score is not None else "",
            score.historic_score if score.historic_score is not None else "",
            score.rarity_score if score.rarity_score is not None else "",
            score.scale_score if score.scale_score is not None else "",
            score.integrity_score if score.integrity_score is not None else "",
            score.fame_score if score.fame_score is not None else "",
            score.season_score if score.season_score is not None else "",
            score.eco_score if score.eco_score is not None else "",
            score.total_score if score.total_score is not None else "",
            score.notes or "", score.scored_by or "",
            score.created_at.isoformat() if score.created_at else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=scores_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


# 问卷构念中文标签（用于导出列名）
QUESTIONNAIRE_CONSTRUCT_LABELS = {
    "trust": "信任",
    "autonomy": "自主",
    "satisfaction": "满意",
    "task_load": "负荷",
    "future_use": "未来意愿",
    "manipulation_check": "操纵检验",
}


async def _compute_questionnaire_scores(db: AsyncSession) -> dict:
    """Aggregate questionnaire responses per user.

    Returns: { user_id: { construct: avg, '_overall': overall_avg,
                           'manipulation_check': raw_text } }.

    Likert (numeric) responses are averaged per construct. The manipulation_check
    construct is *categorical* (choice questions like "是否使用AI助理"), so its raw
    text answers are preserved verbatim (joined by " | ") instead of being averaged —
    otherwise the column would always be empty. This is the key condition-comprehension
    check for the single-blind design.
    """
    result = await db.execute(
        select(
            QuestionnaireResponse.user_id,
            QuestionnaireItem.construct,
            QuestionnaireResponse.response_value,
        ).join(QuestionnaireItem, QuestionnaireResponse.item_id == QuestionnaireItem.id)
    )
    grouped: dict = defaultdict(lambda: defaultdict(list))   # numeric Likert means
    manip: dict = defaultdict(list)                          # raw text for manipulation_check
    for user_id, construct, value in result.all():
        uid = str(user_id)  # key as str to match str(user.id) lookups at export time
        if construct == "manipulation_check":
            text = (str(value).strip() if value is not None else "")
            if text:
                manip[uid].append(text)
            continue
        try:
            num = int(str(value).strip())
        except (ValueError, TypeError):
            continue
        grouped[uid][construct].append(num)

    out: dict = {}
    # Users with Likert responses
    for uid, constructs in grouped.items():
        per = {c: round(sum(v) / len(v), 2) for c, v in constructs.items()}
        all_vals = [x for v in constructs.values() for x in v]
        per["_overall"] = round(sum(all_vals) / len(all_vals), 2) if all_vals else None
        if manip.get(uid):
            per["manipulation_check"] = " | ".join(manip[uid])
        out[uid] = per
    # Users who only answered the manipulation-check questions (no Likert)
    for uid, vals in manip.items():
        if uid not in out:
            out[uid] = {"_overall": None, "manipulation_check": " | ".join(vals)}
    return out


@router.get("/export/all")
async def export_all(
    group: str | None = Query(None),
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export a combined CSV joining participants + submissions + scores + questionnaire.

    Supports `group` filter (H / SOA / MOA) for per-group SPSS export (需求5.3).
    """
    query = (
        select(User, TaskSubmission, AdminScore)
        .outerjoin(TaskSubmission, TaskSubmission.user_id == User.id)
        .outerjoin(AdminScore, AdminScore.user_id == User.id)
        .order_by(User.created_at)
    )
    if group:
        query = query.where(cast(User.group, String) == group)
    rows = (await db.execute(query)).all()

    q_scores = await _compute_questionnaire_scores(db)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Email", "Group", "Status", "Age", "Gender", "Education",
        "Tech_Frequency", "AI_Experience", "Created_At",
        "Task1_Search", "Task2_Document", "Task3_Reminder", "Task4_Email",
        "Duration_Ms", "Submitted_At",
        "Scenic", "Historic", "Rarity", "Scale", "Integrity", "Fame", "Season",
        "Eco", "Total_Score", "Quality_Score", "Reminder_Correct", "Notes",
        "Questionnaire_Overall", "信任", "自主", "满意", "负荷", "未来意愿", "操纵检验",
    ])
    for user, sub, score in rows:
        qs = q_scores.get(str(user.id), {})
        writer.writerow([
            user.email,
            user.group.value if user.group else "",
            user.status.value,
            user.age or "", user.gender or "", user.education or "",
            user.tech_frequency or "", user.ai_experience or "",
            user.created_at.isoformat() if user.created_at else "",
            sub.task1_search if sub else "",
            sub.task2_document if sub else "",
            sub.task3_reminder if sub else "",
            sub.task4_email if sub else "",
            sub.duration_ms if sub else "",
            sub.submitted_at.isoformat() if sub and sub.submitted_at else "",
            score.scenic_score if score and score.scenic_score is not None else "",
            score.historic_score if score and score.historic_score is not None else "",
            score.rarity_score if score and score.rarity_score is not None else "",
            score.scale_score if score and score.scale_score is not None else "",
            score.integrity_score if score and score.integrity_score is not None else "",
            score.fame_score if score and score.fame_score is not None else "",
            score.season_score if score and score.season_score is not None else "",
            score.eco_score if score and score.eco_score is not None else "",
            score.total_score if score and score.total_score is not None else "",
            score.quality_score if score and score.quality_score is not None else "",
            score.reminder_correct if score else "",
            (score.notes or "") if score else "",
            qs.get("_overall", "") if qs else "",
            qs.get("trust", "") if qs else "",
            qs.get("autonomy", "") if qs else "",
            qs.get("satisfaction", "") if qs else "",
            qs.get("task_load", "") if qs else "",
            qs.get("future_use", "") if qs else "",
            qs.get("manipulation_check", "") if qs else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=experiment_all_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


@router.patch("/participants/{user_id}")
async def update_participant_group(
    user_id: str,
    req: dict,
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Manually adjust a participant's experiment group (research balancing / fix mis-assignment).

    Note: the change takes effect on the participant's *next login* — the group is embedded
    in the JWT, so an already-issued token keeps its old group until re-authentication.
    """
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="参与者不存在")

    new_group = req.get("group")
    valid = {g.value for g in ExperimentGroup}
    if new_group not in valid:
        raise HTTPException(status_code=400, detail=f"无效的分组，可选: {sorted(valid)}")

    user.group = ExperimentGroup(new_group)
    await db.commit()
    return {"status": "ok", "user_id": user_id, "group": new_group}


@router.get("/export/all/xlsx")
async def export_all_xlsx(
    group: str | None = Query(None),
    admin: dict = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export experiment data as Excel (.xlsx) with three sheets: 汇总 / 行为日志 / 评分.

    Supports `group` filter (H / SOA / MOA) and includes questionnaire scores (需求5.3).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    def style_header(ws):
        fill = PatternFill("solid", fgColor="1F2937")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Sheet 1: 汇总 (participants + submissions + scores + questionnaire) ----
    combined_q = (
        select(User, TaskSubmission, AdminScore)
        .outerjoin(TaskSubmission, TaskSubmission.user_id == User.id)
        .outerjoin(AdminScore, AdminScore.user_id == User.id)
        .order_by(User.created_at)
    )
    if group:
        combined_q = combined_q.where(cast(User.group, String) == group)
    combined = await db.execute(combined_q)
    q_scores = await _compute_questionnaire_scores(db)

    ws1 = wb.active
    ws1.title = "汇总"
    ws1.append([
        "Email", "Group", "Status", "Age", "Gender", "Education", "Tech_Frequency", "AI_Experience",
        "Created_At", "Task1_Search", "Task2_Document", "Task3_Reminder", "Task4_Email",
        "Duration_Ms", "Submitted_At",
        "Scenic", "Historic", "Rarity", "Scale", "Integrity", "Fame", "Season", "Eco",
        "Total_Score", "Quality_Score", "Reminder_Correct", "Notes",
        "Questionnaire_Overall", "信任", "自主", "满意", "负荷", "未来意愿", "操纵检验",
    ])
    for user, sub, score in combined.all():
        qs = q_scores.get(str(user.id), {})
        ws1.append([
            user.email,
            user.group.value if user.group else "",
            user.status.value,
            user.age or "", user.gender or "", user.education or "",
            user.tech_frequency or "", user.ai_experience or "",
            user.created_at.isoformat() if user.created_at else "",
            sub.task1_search if sub else "",
            sub.task2_document if sub else "",
            sub.task3_reminder if sub else "",
            sub.task4_email if sub else "",
            sub.duration_ms if sub and sub.duration_ms is not None else "",
            sub.submitted_at.isoformat() if sub and sub.submitted_at else "",
            score.scenic_score if score and score.scenic_score is not None else "",
            score.historic_score if score and score.historic_score is not None else "",
            score.rarity_score if score and score.rarity_score is not None else "",
            score.scale_score if score and score.scale_score is not None else "",
            score.integrity_score if score and score.integrity_score is not None else "",
            score.fame_score if score and score.fame_score is not None else "",
            score.season_score if score and score.season_score is not None else "",
            score.eco_score if score and score.eco_score is not None else "",
            score.total_score if score and score.total_score is not None else "",
            score.quality_score if score and score.quality_score is not None else "",
            score.reminder_correct if score else "",
            (score.notes or "") if score else "",
            qs.get("_overall", "") if qs else "",
            qs.get("trust", "") if qs else "",
            qs.get("autonomy", "") if qs else "",
            qs.get("satisfaction", "") if qs else "",
            qs.get("task_load", "") if qs else "",
            qs.get("future_use", "") if qs else "",
            qs.get("manipulation_check", "") if qs else "",
        ])
    style_header(ws1)

    # ---- Sheet 2: 行为日志 ----
    log_rows = (await db.execute(
        select(BehaviorLog, User)
        .join(User, BehaviorLog.user_id == User.id)
        .order_by(BehaviorLog.timestamp)
    )).all()
    ws2 = wb.create_sheet("行为日志")
    ws2.append([
        "Log_ID", "User_ID", "Email", "Group", "Timestamp", "Action_Type", "Action_Target",
        "Input_Content", "Page_Path", "Agent_ID", "Latency_Ms", "Is_Success", "Error_Detail",
        "Phase", "User_Action_On_AI", "Manual_Edit_Count", "Results_Viewed",
        "Clicked_Item_ID", "AI_Suggestion_ID", "AI_Suggestion_Type", "AI_Interaction_Rounds",
        "Final_Plan_Submit_Time",
    ])
    for log, user in log_rows:
        ws2.append([
            str(log.id), str(log.user_id), user.email,
            log.group or "", log.timestamp.isoformat() if log.timestamp else "",
            log.action_type.value, (log.action_target or "")[:200], (log.input_content or "")[:200],
            log.page_path or "", log.agent_id or "",
            log.request_latency_ms if log.request_latency_ms is not None else "",
            log.is_success if log.is_success is not None else "",
            (log.error_detail or "")[:200], log.phase or "",
            log.user_action_on_ai or "",
            log.manual_edit_count if log.manual_edit_count is not None else "",
            log.results_viewed if log.results_viewed is not None else "",
            log.clicked_item_id or "", log.ai_suggestion_id or "",
            log.ai_suggestion_type or "",
            log.ai_interaction_rounds if log.ai_interaction_rounds is not None else "",
            log.final_plan_submit_time.isoformat() if log.final_plan_submit_time else "",
        ])
    style_header(ws2)

    # ---- Sheet 3: 评分 ----
    score_rows = (await db.execute(
        select(AdminScore, User)
        .join(User, AdminScore.user_id == User.id)
    )).all()
    ws3 = wb.create_sheet("评分")
    ws3.append([
        "Score_ID", "User_ID", "Email", "Group",
        "Scenic", "Historic", "Rarity", "Scale", "Integrity", "Fame", "Season",
        "Eco", "Total_Score", "Notes", "Scored_By", "Created_At",
    ])
    for score, user in score_rows:
        ws3.append([
            str(score.id), str(score.user_id), user.email,
            user.group.value if user.group else "",
            score.scenic_score if score.scenic_score is not None else "",
            score.historic_score if score.historic_score is not None else "",
            score.rarity_score if score.rarity_score is not None else "",
            score.scale_score if score.scale_score is not None else "",
            score.integrity_score if score.integrity_score is not None else "",
            score.fame_score if score.fame_score is not None else "",
            score.season_score if score.season_score is not None else "",
            score.eco_score if score.eco_score is not None else "",
            score.total_score if score.total_score is not None else "",
            score.notes or "", score.scored_by or "",
            score.created_at.isoformat() if score.created_at else "",
        ])
    style_header(ws3)

    # Auto-size-ish: set a reasonable width for readability
    for ws in (ws1, ws2, ws3):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=experiment_data_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )
